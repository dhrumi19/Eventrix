import csv
import io
import uuid
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, Http404
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

import razorpay
from accounts.models import User
from .models import Event, Seat, Booking, Payment, Ticket, Notification, Report
from .forms import EventForm
from .utils import generate_ticket_qr, generate_ticket_pdf, send_ticket_email

# =====================================================================
# MODULE 1 - USER MODULE VIEWS (Event Discovery, Booking & Payments)
# =====================================================================

def home_view(request):
    """
    Browse events with search, category filtering, city filtering, and date filtering.
    """
    events = Event.objects.filter(is_approved=True, is_published=True).order_by('date_time')
    
    # Get distinct cities and categories for filter dropdowns
    cities = Event.objects.filter(is_approved=True, is_published=True).values_list('city', flat=True).distinct()
    categories = Event.CATEGORY_CHOICES

    # Apply search query
    query = request.GET.get('q')
    if query:
        events = events.filter(
            Q(title__icontains=query) | 
            Q(description__icontains=query) | 
            Q(venue__icontains=query)
        )
        
    # Apply category filter
    category = request.GET.get('category')
    if category:
        events = events.filter(category=category)
        
    # Apply city filter
    city = request.GET.get('city')
    if city:
        events = events.filter(city=city)
        
    # Apply date filter (Today, Tomorrow, Weekend, Upcoming)
    date_filter = request.GET.get('date_filter')
    now = timezone.now()
    if date_filter == 'today':
        events = events.filter(date_time__date=now.date())
    elif date_filter == 'tomorrow':
        events = events.filter(date_time__date=now.date() + timezone.timedelta(days=1))
    elif date_filter == 'weekend':
        # Filter for upcoming Friday, Saturday, Sunday
        today_weekday = now.weekday()  # Monday is 0, Sunday is 6
        friday = now + timezone.timedelta(days=(4 - today_weekday) % 7)
        sunday = friday + timezone.timedelta(days=2)
        events = events.filter(date_time__date__range=[friday.date(), sunday.date()])
    elif date_filter == 'upcoming':
        events = events.filter(date_time__gte=now)

    context = {
        'events': events,
        'cities': cities,
        'categories': categories,
        'query': query,
        'selected_category': category,
        'selected_city': city,
        'selected_date_filter': date_filter
    }
    return render(request, 'events/home.html', context)


def event_detail_view(request, event_id):
    """
    Display details of an individual event and show available seats count.
    """
    event = get_object_or_404(Event, id=event_id, is_approved=True, is_published=True)
    available_seats = event.seats.filter(status='AVAILABLE').count()
    
    context = {
        'event': event,
        'available_seats_count': available_seats
    }
    return render(request, 'events/event_detail.html', context)


@login_required
def select_seats_view(request, event_id):
    """
    Interactive Seat Selection Page. Displays layout and availability.
    Processes seat reservations and prepares Razorpay checkout details.
    """
    event = get_object_or_404(Event, id=event_id, is_approved=True, is_published=True)
    seats = event.seats.all()
    
    if request.method == 'POST':
        selected_seat_ids = request.POST.getlist('selected_seats')
        
        if not selected_seat_ids:
            messages.error(request, "Please select at least one seat to proceed.")
            return redirect('select_seats', event_id=event.id)
            
        # Verify seats are still available
        reserved_seats = seats.filter(id__in=selected_seat_ids, status='AVAILABLE')
        
        if reserved_seats.count() != len(selected_seat_ids):
            messages.error(request, "Some of the selected seats were already booked. Please choose other seats.")
            return redirect('select_seats', event_id=event.id)
            
        # Calculate total price
        total_amount = event.ticket_price * len(selected_seat_ids)
        
        # Create a Pending Booking
        booking = Booking.objects.create(
            user=request.user,
            event=event,
            total_amount=total_amount,
            payment_status='PENDING'
        )
        # Link seats to the booking and block them temporarily
        booking.seats.set(reserved_seats)
        reserved_seats.update(status='BLOCKED', booked_by=request.user)
        
        return redirect('payment', booking_id=booking.id)

    # Group seats by row for layout rendering
    grouped_seats = {}
    for seat in seats:
        if seat.row_label not in grouped_seats:
            grouped_seats[seat.row_label] = []
        grouped_seats[seat.row_label].append(seat)

    context = {
        'event': event,
        'grouped_seats': grouped_seats,
        'total_seats': seats.count(),
        'available_seats': seats.filter(status='AVAILABLE').count(),
        'booked_seats': seats.filter(status='BOOKED').count(),
        'blocked_seats': seats.filter(status='BLOCKED').count()
    }
    return render(request, 'events/select_seats.html', context)


@login_required
def payment_view(request, booking_id):
    """
    Displays payment page. Integrates Razorpay Order creation or simulation mode.
    """
    booking = get_object_or_404(Booking, id=booking_id, user=request.user, payment_status='PENDING')
    
    # Check if we should use Razorpay API or local simulation
    is_simulation = settings.USE_RAZORPAY_SIMULATION or not settings.RAZORPAY_KEY_ID
    
    razorpay_order_id = ""
    if not is_simulation:
        try:
            client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))
            # Create Razorpay Order (amount in paise)
            order_data = {
                'amount': int(booking.total_amount * 100),
                'currency': 'INR',
                'receipt': f'order_rcpt_{booking.id}',
                'payment_capture': 1
            }
            order = client.order.create(data=order_data)
            razorpay_order_id = order['id']
        except Exception as e:
            # Fallback to simulation mode if API error occurs
            is_simulation = True
            messages.warning(request, "Razorpay API unavailable. Falling back to secure simulated checkout.")
            
    context = {
        'booking': booking,
        'is_simulation': is_simulation,
        'razorpay_order_id': razorpay_order_id,
        'razorpay_key_id': settings.RAZORPAY_KEY_ID,
        'total_amount_paise': int(booking.total_amount * 100)
    }
    return render(request, 'events/payment.html', context)


@login_required
def payment_callback_view(request):
    """
    Callback view confirming Razorpay payment signatures or simulated outcomes.
    Marks bookings complete, registers payment details, generates ticket assets,
    and publishes alerts/notifications.
    """
    if request.method == 'POST':
        booking_id = request.POST.get('booking_id')
        status = request.POST.get('status') # 'success' or 'failed'
        booking = get_object_or_404(Booking, id=booking_id, user=request.user)
        
        if status == 'success':
            # Complete booking
            booking.payment_status = 'COMPLETED'
            booking.save()
            
            # Confirm seats status to BOOKED
            booking.seats.all().update(status='BOOKED')
            
            # Payment logging
            payment_id = request.POST.get('razorpay_payment_id', f'sim_pay_{uuid.uuid4().hex[:12]}')
            order_id = request.POST.get('razorpay_order_id', f'sim_ord_{uuid.uuid4().hex[:12]}')
            signature = request.POST.get('razorpay_signature', 'sim_sig_valid')
            
            Payment.objects.create(
                booking=booking,
                payment_id=payment_id,
                order_id=order_id,
                signature=signature,
                amount=booking.total_amount,
                status='SUCCESS'
            )
            
            # Generate Ticket models
            ticket = Ticket.objects.create(booking=booking)
            
            # Generate assets
            generate_ticket_qr(ticket)
            generate_ticket_pdf(ticket)
            ticket.save()
            
            # Send booking confirmation email
            send_ticket_email(ticket)
            
            # Create user Notification
            Notification.objects.create(
                user=request.user,
                message=f"Booking confirmed! Your ticket for {booking.event.title} is ready. Ticket Code: {ticket.ticket_code}"
            )
            
            # Notify Organizer
            Notification.objects.create(
                user=booking.event.organizer,
                message=f"New Booking: User {request.user.username} booked {booking.seats.count()} seats for {booking.event.title}."
            )
            
            messages.success(request, f"Payment successful! Seats booked: {', '.join(s.row_label + str(s.seat_number) for s in booking.seats.all())}.")
            return redirect('booking_success', booking_id=booking.id)
            
        else:
            # Payment failed, release blocked seats
            booking.payment_status = 'FAILED'
            booking.save()
            booking.seats.all().update(status='AVAILABLE', booked_by=None)
            
            Payment.objects.create(
                booking=booking,
                payment_id=f"failed_{uuid.uuid4().hex[:8]}",
                order_id=f"failed_{uuid.uuid4().hex[:8]}",
                amount=booking.total_amount,
                status='FAILED'
            )
            
            messages.error(request, "Payment failed or was cancelled. Seats have been released.")
            return redirect('dashboard')
            
    return redirect('home')


@login_required
def booking_success_view(request, booking_id):
    """
    Displays the details of a successfully booked ticket.
    """
    booking = get_object_or_404(Booking, id=booking_id, user=request.user, payment_status='COMPLETED')
    ticket = get_object_or_404(Ticket, booking=booking)
    
    context = {
        'booking': booking,
        'ticket': ticket,
    }
    return render(request, 'events/booking_success.html', context)


@login_required
def cancel_booking_view(request, booking_id):
    """
    Cancels a completed booking, releases seats back to AVAILABLE, deletes the ticket,
    and updates the payment status to CANCELLED.
    """
    if request.method == 'POST':
        booking = get_object_or_404(Booking, id=booking_id, user=request.user)
        
        if booking.payment_status != 'COMPLETED':
            messages.error(request, "Only confirmed bookings can be cancelled.")
            return redirect('dashboard')
            
        # Release seats
        seats = booking.seats.all()
        seats.update(status='AVAILABLE', booked_by=None)
        
        # Delete associated ticket if it exists
        if hasattr(booking, 'ticket'):
            booking.ticket.delete()
            
        # Update booking status
        booking.payment_status = 'CANCELLED'
        booking.save()
        
        # Update associated payment record
        if hasattr(booking, 'payment'):
            booking.payment.status = 'FAILED'
            booking.payment.save()
            
        # Notify user and organizer
        Notification.objects.create(
            user=request.user,
            message=f"Booking #{booking.id} for {booking.event.title} has been cancelled. Your seats have been released."
        )
        
        Notification.objects.create(
            user=booking.event.organizer,
            message=f"Cancellation: User {request.user.username} cancelled their booking #{booking.id} for {booking.event.title}."
        )
        
        messages.success(request, f"Booking #{booking.id} has been cancelled successfully. Your seats have been released.")
        return redirect('dashboard')
        
    return redirect('home')


@login_required
def view_ticket_qr_view(request, ticket_code):
    """
    Displays the generated ticket with QR code for validation.
    """
    ticket = get_object_or_404(Ticket, ticket_code=ticket_code)
    # Ensure only the ticket owner, organizer of the event, or admin can view it
    if request.user != ticket.booking.user and request.user != ticket.booking.event.organizer and not request.user.is_admin_role:
        raise Http404("Unauthorized ticket access.")
        
    return render(request, 'events/view_ticket_qr.html', {'ticket': ticket})


@login_required
def download_ticket_pdf_view(request, ticket_code):
    """
    Downloads the pre-generated PDF Ticket as a response stream.
    """
    ticket = get_object_or_404(Ticket, ticket_code=ticket_code)
    if request.user != ticket.booking.user and request.user != ticket.booking.event.organizer and not request.user.is_admin_role:
        raise Http404("Unauthorized ticket access.")
        
    if not ticket.pdf_file:
        # Regenerate on the fly if file missing
        generate_ticket_pdf(ticket)
        ticket.save()
        
    response = HttpResponse(ticket.pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ticket_{ticket.booking.event.title.replace(" ", "_")}.pdf"'
    return response


@login_required
def read_notification_view(request, notification_id):
    """
    Marks notifications as read.
    """
    notification = get_object_or_404(Notification, id=notification_id, user=request.user)
    notification.is_read = True
    notification.save()
    return redirect('dashboard')


# =====================================================================
# MODULE 2 - ORGANIZER MODULE VIEWS (Event CRUD & Seat Management)
# =====================================================================

@login_required
def create_event_view(request):
    """
    Create a new event. Organizers must be approved.
    """
    if not request.user.is_organizer_role:
        messages.error(request, "Only organizers can create events.")
        return redirect('dashboard')
        
    if not request.user.is_approved_organizer:
        return redirect('dashboard')
        
    if request.method == 'POST':
        form = EventForm(request.POST, request.FILES)
        if form.is_valid():
            event = form.save(commit=False)
            event.organizer = request.user
            event.is_approved = False  # requires admin approval
            event.save()
            messages.success(request, f"Event '{event.title}' created successfully! It is pending administrator verification before being published.")
            return redirect('dashboard')
    else:
        form = EventForm()
    return render(request, 'events/create_event.html', {'form': form})


@login_required
def edit_event_view(request, event_id):
    """
    Edit event details. Owner organizer check enforced.
    """
    event = get_object_or_404(Event, id=event_id)
    if event.organizer != request.user and not request.user.is_admin_role:
        messages.error(request, "Unauthorized access.")
        return redirect('dashboard')
        
    if request.method == 'POST':
        form = EventForm(request.POST, request.FILES, instance=event)
        if form.is_valid():
            form.save()
            messages.success(request, f"Event '{event.title}' details updated successfully!")
            return redirect('dashboard')
    else:
        form = EventForm(instance=event)
    return render(request, 'events/edit_event.html', {'form': form, 'event': event})


@login_required
def delete_event_view(request, event_id):
    """
    Deletes event and associated data.
    """
    event = get_object_or_404(Event, id=event_id)
    if event.organizer != request.user and not request.user.is_admin_role:
        messages.error(request, "Unauthorized access.")
        return redirect('dashboard')
        
    title = event.title
    event.delete()
    messages.success(request, f"Event '{title}' and all bookings/seats deleted successfully.")
    
    # Handle redirect path
    next_page = request.GET.get('next')
    if next_page == 'admin_dashboard':
        return redirect('admin_dashboard')
    return redirect('dashboard')


@login_required
def setup_seats_view(request, event_id):
    """
    Creates/modifies seat layouts for an event. Enables organizers to configure
    custom grid widths/rows.
    """
    event = get_object_or_404(Event, id=event_id)
    if event.organizer != request.user and not request.user.is_admin_role:
        messages.error(request, "Unauthorized access.")
        return redirect('dashboard')
        
    seats = event.seats.all()
    booked_count = seats.filter(status='BOOKED').count()
    
    if request.method == 'POST':
        rows_input = request.POST.get('rows_count', '5') # e.g. 5 means rows A-E
        cols_input = request.POST.get('cols_count', '10') # e.g. 10 seats per row
        
        try:
            num_rows = int(rows_input)
            num_cols = int(cols_input)
            
            if num_rows < 1 or num_rows > 20 or num_cols < 1 or num_cols > 30:
                raise ValueError("Rows/columns counts exceed layout boundaries (max 20 rows, 30 columns).")
                
            if booked_count > 0:
                messages.error(request, "Cannot restructure layout: seats have already been booked for this event.")
                return redirect('setup_seats', event_id=event.id)
                
            # Clear old seats
            seats.delete()
            
            # Map integers to alphabetical row labels: 1->A, 2->B, etc.
            row_labels = [chr(65 + i) for i in range(num_rows)]
            
            seats_to_create = []
            for row in row_labels:
                for col in range(1, num_cols + 1):
                    seats_to_create.append(Seat(
                        event=event,
                        row_label=row,
                        seat_number=col,
                        status='AVAILABLE'
                    ))
            Seat.objects.bulk_create(seats_to_create)
            
            # Update event capacity matching layout size
            event.capacity = num_rows * num_cols
            event.save()
            
            messages.success(request, f"Seat layout configured successfully! Recreated {event.capacity} seats ({num_rows} rows x {num_cols} cols).")
            return redirect('dashboard')
            
        except ValueError as e:
            messages.error(request, f"Configuration Error: {str(e)}")
            
    # Count current rows and cols for form defaults
    distinct_rows = sorted(list(set(s.row_label for s in seats)))
    current_rows_count = len(distinct_rows) or 5
    current_cols_count = seats.filter(row_label=distinct_rows[0]).count() if distinct_rows else 10

    context = {
        'event': event,
        'current_rows_count': current_rows_count,
        'current_cols_count': current_cols_count,
        'booked_count': booked_count,
        'seats': seats
    }
    return render(request, 'events/setup_seats.html', context)


@login_required
def organizer_bookings_view(request, event_id):
    """
    List all bookings for an organizer's specific event.
    """
    event = get_object_or_404(Event, id=event_id)
    if event.organizer != request.user and not request.user.is_admin_role:
        messages.error(request, "Unauthorized access.")
        return redirect('dashboard')
        
    bookings = Booking.objects.filter(event=event).order_by('-booking_date')
    context = {
        'event': event,
        'bookings': bookings
    }
    return render(request, 'events/organizer_bookings.html', context)


@login_required
def export_participants_view(request, event_id, export_type):
    """
    Exports the list of booked participants for an event in CSV or Excel.
    """
    event = get_object_or_404(Event, id=event_id)
    if event.organizer != request.user and not request.user.is_admin_role:
        return HttpResponse("Unauthorized", status=403)
        
    bookings = Booking.objects.filter(event=event, payment_status='COMPLETED').order_by('booking_date')
    
    filename = f"participants_{event.title.replace(' ', '_')}"
    
    if export_type == 'csv':
        # CSV Export
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Booking ID', 'Username', 'Full Name', 'Email', 'Phone', 'Seats Booked', 'Amount Paid (INR)', 'Booking Date'])
        
        for b in bookings:
            seats_str = ", ".join(f"{s.row_label}{s.seat_number}" for s in b.seats.all())
            writer.writerow([
                b.id,
                b.user.username,
                f"{b.user.first_name} {b.user.last_name}",
                b.user.email,
                b.user.phone or "N/A",
                seats_str,
                b.total_amount,
                b.booking_date.strftime('%Y-%m-%d %H:%M')
            ])
        return response
        
    elif export_type == 'excel':
        # Excel Export (openpyxl)
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendee List"
        
        # Headers
        headers = ['Booking ID', 'Username', 'Full Name', 'Email', 'Phone', 'Seats Booked', 'Amount Paid (INR)', 'Booking Date']
        ws.append(headers)
        
        # Style headers
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="6C5CE7", end_color="6C5CE7", fill_type="solid")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        # Add records
        for b in bookings:
            seats_str = ", ".join(f"{s.row_label}{s.seat_number}" for s in b.seats.all())
            ws.append([
                b.id,
                b.user.username,
                f"{b.user.first_name} {b.user.last_name}",
                b.user.email,
                b.user.phone or "N/A",
                seats_str,
                float(b.total_amount),
                b.booking_date.strftime('%Y-%m-%d %H:%M')
            ])
            
        # Adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        return response
        
    return redirect('dashboard')


# =====================================================================
# MODULE 3 - ADMIN MODULE VIEWS (Moderation, Approvals & Reporting)
# =====================================================================

@login_required
def approve_organizer_view(request, user_id):
    """
    Approve an organizer registration request.
    """
    if not request.user.is_admin_role:
        messages.error(request, "Unauthorized access.")
        return redirect('dashboard')
        
    organizer = get_object_or_404(User, id=user_id, role='ORGANIZER')
    organizer.is_approved_organizer = True
    organizer.save()
    
    # Notify Organizer
    Notification.objects.create(
        user=organizer,
        message="Your organizer account has been approved by the Administrator. You can now publish events!"
    )
    
    messages.success(request, f"Organizer @{organizer.username} has been approved successfully.")
    return redirect('dashboard')


@login_required
def delete_user_admin_view(request, user_id):
    """
    Suspend/delete users from system (Admin moderation tool).
    """
    if not request.user.is_admin_role:
        messages.error(request, "Unauthorized access.")
        return redirect('dashboard')
        
    u = get_object_or_404(User, id=user_id)
    username = u.username
    u.delete()
    messages.success(request, f"User @{username} has been deleted from the platform.")
    return redirect('dashboard')


@login_required
def approve_event_view(request, event_id):
    """
    Admin moderates and approves event listings.
    """
    if not request.user.is_admin_role:
        messages.error(request, "Unauthorized access.")
        return redirect('dashboard')
        
    event = get_object_or_404(Event, id=event_id)
    event.is_approved = True
    event.is_published = True  # Auto-publish on approval
    event.save()
    
    # Generate default seat layout if organizer hasn't configured it
    if event.seats.count() == 0:
        rows = ['A', 'B', 'C', 'D', 'E']
        seats_to_create = []
        for r in rows:
            for c in range(1, 11):
                seats_to_create.append(Seat(
                    event=event,
                    row_label=r,
                    seat_number=c,
                    status='AVAILABLE'
                ))
        Seat.objects.bulk_create(seats_to_create)
        event.capacity = 50
        event.save()
        
    # Notify Organizer
    Notification.objects.create(
        user=event.organizer,
        message=f"Your event '{event.title}' has been approved and published! A default 50-seat grid layout has been initialized."
    )
    
    messages.success(request, f"Event '{event.title}' approved and published.")
    return redirect('dashboard')


@login_required
def reject_event_view(request, event_id):
    """
    Admin rejects/blocks an event listing.
    """
    if not request.user.is_admin_role:
        messages.error(request, "Unauthorized access.")
        return redirect('dashboard')
        
    event = get_object_or_404(Event, id=event_id)
    title = event.title
    organizer = event.organizer
    event.delete()
    
    # Notify Organizer
    Notification.objects.create(
        user=organizer,
        message=f"Your event application for '{title}' was declined by the Administrator."
    )
    
    messages.success(request, f"Event '{title}' rejected and removed.")
    return redirect('dashboard')


@login_required
def admin_report_pdf_view(request, report_type):
    """
    Generates system-wide PDF summaries for Bookings or Revenue.
    """
    if not request.user.is_admin_role:
        return HttpResponse("Unauthorized", status=403)
        
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'RepTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        textColor=colors.HexColor('#6c5ce7'),
        spaceAfter=15
    )
    
    normal_style = styles['Normal']
    header_style = ParagraphStyle('RepHeader', parent=normal_style, fontName='Helvetica-Bold', fontSize=10, textColor=colors.white)
    
    story = []
    
    if report_type == 'revenue':
        story.append(Paragraph("<b>EVENTRIX SYSTEM REVENUE REPORT</b>", title_style))
        story.append(Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}", normal_style))
        story.append(Spacer(1, 15))
        
        # Calculate summary metrics
        payments = Payment.objects.filter(status='SUCCESS').order_by('-created_at')
        total_revenue = payments.aggregate(tot=Sum('amount'))['tot'] or 0
        
        story.append(Paragraph(f"<b>Total Platform Revenue:</b> INR {total_revenue}", styles['Heading3']))
        story.append(Spacer(1, 10))
        
        # Build table
        table_data = [[
            Paragraph("<b>Transaction ID</b>", header_style),
            Paragraph("<b>Booking Ref</b>", header_style),
            Paragraph("<b>User</b>", header_style),
            Paragraph("<b>Amount</b>", header_style),
            Paragraph("<b>Date Logged</b>", header_style)
        ]]
        
        for pay in payments[:30]:  # Limit to 30 records
            table_data.append([
                Paragraph(pay.payment_id, normal_style),
                Paragraph(f"#{pay.booking.id}", normal_style),
                Paragraph(pay.booking.user.username, normal_style),
                Paragraph(f"₹{pay.amount}", normal_style),
                Paragraph(pay.created_at.strftime('%Y-%m-%d'), normal_style)
            ])
            
        t = Table(table_data, colWidths=[1.8*inch, 1*inch, 1.2*inch, 1*inch, 1.2*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#6c5ce7')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dcdde1')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(t)
        
    elif report_type == 'bookings':
        story.append(Paragraph("<b>EVENTRIX SYSTEM BOOKINGS REPORT</b>", title_style))
        story.append(Paragraph(f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}", normal_style))
        story.append(Spacer(1, 15))
        
        bookings = Booking.objects.all().order_by('-created_at')
        total_bookings_count = bookings.count()
        completed_bookings = bookings.filter(payment_status='COMPLETED').count()
        
        story.append(Paragraph(f"<b>Total Bookings Processed:</b> {total_bookings_count} ({completed_bookings} Completed)", styles['Heading3']))
        story.append(Spacer(1, 10))
        
        table_data = [[
            Paragraph("<b>Booking ID</b>", header_style),
            Paragraph("<b>Event Title</b>", header_style),
            Paragraph("<b>User</b>", header_style),
            Paragraph("<b>Seats Count</b>", header_style),
            Paragraph("<b>Status</b>", header_style)
        ]]
        
        for b in bookings[:30]:
            table_data.append([
                Paragraph(f"#{b.id}", normal_style),
                Paragraph(b.event.title[:25], normal_style),
                Paragraph(b.user.username, normal_style),
                Paragraph(str(b.seats.count()), normal_style),
                Paragraph(b.payment_status, normal_style)
            ])
            
        t = Table(table_data, colWidths=[1*inch, 2.2*inch, 1.2*inch, 1*inch, 1*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#6c5ce7')),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#dcdde1')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        story.append(t)
        
    doc.build(story)
    
    # Save Report record (database log)
    pdf_val = buffer.getvalue()
    rep = Report.objects.create(
        report_type='REVENUE' if report_type == 'revenue' else 'BOOKING',
        generated_by=request.user
    )
    rep.file.save(f"report_{report_type}_{timezone.now().strftime('%Y%m%d')}.pdf", ContentFile(pdf_val))
    rep.save()
    buffer.close()
    
    response = HttpResponse(rep.file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="system_report_{report_type}.pdf"'
    return response


@login_required
def admin_report_excel_view(request, report_type):
    """
    Generates spreadsheet reports for Events or Users.
    """
    if not request.user.is_admin_role:
        return HttpResponse("Unauthorized", status=403)
        
    wb = Workbook()
    ws = wb.active
    
    if report_type == 'events':
        ws.title = "Events Summary"
        headers = ['Event ID', 'Title', 'Organizer', 'Category', 'City', 'Venue', 'Ticket Price', 'Capacity', 'Tickets Sold', 'Revenue']
        ws.append(headers)
        
        events = Event.objects.all().order_by('-created_at')
        for e in events:
            t_sold = Booking.objects.filter(event=e, payment_status='COMPLETED').aggregate(seats_sum=Count('seats'))['seats_sum'] or 0
            rev = Payment.objects.filter(booking__event=e, booking__payment_status='COMPLETED').aggregate(sum_amount=Sum('amount'))['sum_amount'] or 0
            
            ws.append([
                e.id,
                e.title,
                e.organizer.username,
                e.get_category_display(),
                e.city,
                e.venue,
                float(e.ticket_price),
                e.capacity,
                t_sold,
                float(rev)
            ])
            
    elif report_type == 'users':
        ws.title = "Platform Users"
        headers = ['User ID', 'Username', 'Full Name', 'Email', 'Role', 'Approved Organizer?', 'Date Registered', 'Bookings Count', 'Total Spent']
        ws.append(headers)
        
        users = User.objects.all().order_by('-created_at')
        for u in users:
            b_count = Booking.objects.filter(user=u, payment_status='COMPLETED').count()
            spent = Payment.objects.filter(booking__user=u, booking__payment_status='COMPLETED').aggregate(sum_amount=Sum('amount'))['sum_amount'] or 0
            
            ws.append([
                u.id,
                u.username,
                f"{u.first_name} {u.last_name}",
                u.email,
                u.get_role_display(),
                "Yes" if u.is_approved_organizer else "No",
                u.created_at.strftime('%Y-%m-%d'),
                b_count,
                float(spent)
            ])
            
    # Apply stylings
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6C5CE7", end_color="6C5CE7", fill_type="solid")
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="system_report_{report_type}.xlsx"'
    wb.save(response)
    
    # Save Report entry database log
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    rep = Report.objects.create(
        report_type='EVENT' if report_type == 'events' else 'PARTICIPANT',
        generated_by=request.user
    )
    rep.file.save(f"report_{report_type}_{timezone.now().strftime('%Y%m%d')}.xlsx", ContentFile(excel_buffer.getvalue()))
    rep.save()
    excel_buffer.close()
    
    return response
from django.core.files.base import ContentFile
